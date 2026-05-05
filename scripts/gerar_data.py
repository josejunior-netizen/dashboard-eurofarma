#!/usr/bin/env python3
"""
gerar_data.py — Eurofarma Dashboard
Baixa o XLS de OS do OneDrive/SharePoint, cruza com Sourcing 2026
e Vol. Hotelaria, e gera o arquivo data.js que o HTML consome.
"""

import os
import sys
import json
import unicodedata
import re
import requests
from collections import defaultdict, Counter
from datetime import datetime, timezone, timedelta
import openpyxl

ONEDRIVE_URL = os.environ.get("ONEDRIVE_URL", "")
SOURCING_URL = os.environ.get("SOURCING_URL", "")
OUTPUT_FILE  = "data.js"
DIV_THRESHOLD = 0.12
TZ_BR = timezone(timedelta(hours=-3))

def norm(s):
    s = unicodedata.normalize("NFD", str(s).upper().strip())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^A-Z0-9 /]", "", s).strip()

def safe_float(v):
    try:
        f = float(str(v).replace(",", "."))
        return f if f > 0 else None
    except Exception:
        return None

def fmt_date(val):
    if not val or str(val).strip() in ("", "None", "nan"):
        return ""
    try:
        if isinstance(val, datetime):
            return val.strftime("%d/%m")
        dt = datetime.fromisoformat(str(val)[:10])
        return dt.strftime("%d/%m")
    except Exception:
        return str(val)[:5]

def download_bytes(url, label):
    """
    Retorna (bytes, last_modified_dt).
    last_modified_dt é o datetime do header HTTP Last-Modified, ou None se ausente.
    OneDrive/SharePoint expõem Last-Modified com a hora real da última edição do arquivo.
    """
    print(f"  Baixando {label}...")
    resp = requests.get(url, timeout=60)
    resp.raise_for_status()

    last_modified_dt = None
    last_modified_header = resp.headers.get("Last-Modified", "")
    if last_modified_header:
        try:
            from email.utils import parsedate_to_datetime
            last_modified_dt = parsedate_to_datetime(last_modified_header)
            # Converter para fuso de Brasília
            if last_modified_dt.tzinfo:
                last_modified_dt = last_modified_dt.astimezone(TZ_BR)
            print(f"  ✓ {label}: {len(resp.content):,} bytes (Last-Modified: {last_modified_dt.strftime('%d/%m/%Y %H:%M')})")
        except Exception as e:
            print(f"  ⚠ {label}: falha parsear Last-Modified ({e})")
            print(f"  ✓ {label}: {len(resp.content):,} bytes")
    else:
        print(f"  ✓ {label}: {len(resp.content):,} bytes (sem Last-Modified)")

    return resp.content, last_modified_dt

def carregar_sourcing(raw_bytes):
    import io
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True)
    ws = wb["Sourcing 2026 - Eurofarma"]
    rows = list(ws.iter_rows(values_only=True))
    sourcing = {}
    for r in rows[1:]:
        status = str(r[4]).strip() if r[4] else ""
        if status not in ("Sim", "Offline"):
            continue
        cidade = str(r[2]).strip().upper() if r[2] else ""
        hotel  = str(r[3]).strip().upper() if r[3] else ""
        key    = f"{hotel}||{cidade}"
        blackout = str(r[23]).strip() if r[23] else ""
        if blackout.lower() in ("não", "none", "nan", "não, as taxas são válidas para todas as datas.", ""):
            blackout = ""
        email = str(r[25]).strip() if r[25] else ""
        if email.lower() in ("none", "nan", ""):
            email = ""
        sourcing[key] = {
            "hotel":  hotel,
            "cidade": cidade,
            "status": status,
            "tarifa_ind":  safe_float(r[5]),
            "tarifa_dup":  safe_float(r[6]),
            "tipo_tarifa": str(r[10]).strip() if r[10] else "",
            "condicao": ("LRA" if r[11] and "NLRA" not in str(r[11]) and "LRA" in str(r[11]) else "NLRA") if r[11] else "",
            "blackout": blackout[:60],
            "email":   email,
            "cafe":    str(r[19]).strip() if r[19] else "",
        }
    print(f"  ✓ Sourcing: {len(sourcing)} acordos ativos")
    return sourcing

def carregar_historico(raw_bytes):
    """Retorna dicts de emissores, tarifas, formas de pagamento E set de OS finalizadas."""
    import io
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True)
    ws = wb["Vol. Hotelaria 2026"]
    rows = list(ws.iter_rows(values_only=True))
    hotel_hist_sgl  = defaultdict(list)
    hotel_hist_dbl  = defaultdict(list)
    hotel_emissores = defaultdict(Counter)
    hotel_pagamento = defaultdict(Counter)  # ← NOVO: forma de pagamento por hotel
    os_finalizadas = set()  # ← NOVO: números de OS já emitidas (para filtrar pendentes)
    for r in rows[1:]:
        hotel      = str(r[17]).strip().upper() if r[17] else ""
        cidade     = str(r[32]).strip().upper() if r[32] else ""
        emissor    = str(r[65]).strip() if r[65] else ""
        tipo_apto  = str(r[40]).strip().upper() if r[40] else "SGL"
        # Coluna B = índice 1 = Número da OS (toda OS aqui já foi emitida no ARGO)
        num_os_raw = r[1]
        # Coluna X = índice 23 = Forma de Pagamento
        forma_pgto = str(r[23]).strip() if r[23] else ""
        # Registrar OS como finalizada (mesmo se outras colunas estão vazias)
        if num_os_raw is not None and str(num_os_raw).strip() not in ("", "None", "nan"):
            try:
                # Pode vir como int, float ou string com decimal
                num_os = int(float(str(num_os_raw).strip()))
                os_finalizadas.add(num_os)
            except (ValueError, TypeError):
                pass
        if not hotel or emissor in ("None", "COPASTUR", "", "nan"):
            continue
        try:
            valor = float(str(r[48]).replace(",", ".")) if r[48] else None
        except Exception:
            valor = None
        key = f"{hotel}||{cidade}"
        if valor and valor > 0:
            if tipo_apto == "DBL":
                hotel_hist_dbl[key].append(valor)
            else:
                hotel_hist_sgl[key].append(valor)
        hotel_emissores[key][emissor] += 1
        # Registrar forma de pagamento (ignora valores vazios/inválidos)
        if forma_pgto and forma_pgto.lower() not in ("none", "nan", ""):
            hotel_pagamento[key][forma_pgto] += 1
    print(f"  OK Historico: {len(hotel_emissores)} hoteis com reservas, {len(os_finalizadas)} OS finalizadas")
    return hotel_hist_sgl, hotel_hist_dbl, hotel_emissores, hotel_pagamento, os_finalizadas

def buscar_sourcing(hotel_name, cidade_str, sourcing):
    """
    Busca acordo de sourcing para o hotel da OS.
    Regras (mesmas do calcScore JS):
    - Cidade é REQUISITO (não bate, descarta)
    - NOME do hotel é OBRIGATÓRIO (sem match de nome, score=0)
    - Palavras da cidade NÃO contam como nome (ex: VISTA em "Blue Inn Boa Vista")
    - Núcleo igual em qualquer ordem = match exato (ex: PRATA HOTEL ↔ HOTEL PRATA)
    - cidade_compativel: match por prefixo (>=4 chars), não substring arbitrário
    - Threshold dinâmico: 12/8/5 conforme palavras significativas
    """
    hn = norm(hotel_name)
    cn = norm(cidade_str.split("/")[0].strip())

    PALAVRAS_GENERICAS = {'HOTEL','PALACE','POUSADA','APART','PLAZA','PARK',
                          'PREMIUM','INN','EXPRESS','RESORT','FLAT','COMFORT',
                          'QUALITY','GRAND','SUITE','EXECUTIVE','SUITES','SUITE',
                          'PLUS','BEST','PRIME','GRANDE','FLATS'}

    # cidade_compativel — match por prefixo (mín 4 chars), não substring arbitrário
    # Evita "PALMAS" casar com "ALMAS" só porque ALMAS está dentro de PALMAS
    def cidade_compativel(email_c, plan_c):
        if not email_c: return True
        def norm_c(s):
            s = unicodedata.normalize('NFD', s.upper())
            s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
            return ' '.join(s.replace('/', ' ').split())
        eN = norm_c(email_c); pN = norm_c(plan_c)
        if eN == pN: return True
        PREP = {'DO','DA','DE','DI','DOS','DAS','E','O','A'}
        palE = [p for p in eN.split() if len(p) > 1 and p not in PREP]
        palP = [p for p in pN.split() if len(p) > 1 and p not in PREP]
        if abs(len(palE) - len(palP)) > 1: return False
        def palavra_casa(a, b):
            if a == b: return True
            if len(a) < 4 or len(b) < 4: return False
            mn = min(len(a), len(b))
            return a[:mn] == b[:mn]
        menor = palE if len(palE) <= len(palP) else palP
        maior = palP if len(palE) <= len(palP) else palE
        return all(any(palavra_casa(m, p) for m in maior) for p in menor)

    melhor = None
    melhor_score = 0

    for k, v in sourcing.items():
        kh = norm(v["hotel"])
        kc = norm(v["cidade"])

        # 1) Cidade é REQUISITO
        cidade_bate = cidade_compativel(cn, kc)
        if cn and not cidade_bate:
            continue

        # Palavras da cidade não devem contar como nome do hotel
        # Ex: "VISTA" em "Blue Inn Boa Vista" é da cidade Boa Vista
        palavras_cidade = set()
        for c in (cn, kc):
            if c:
                for p in c.split():
                    if len(p) > 2:
                        palavras_cidade.add(p)
        def nao_eh_cidade(p):
            return p not in palavras_cidade

        # 2) NOME do hotel é OBRIGATÓRIO
        score_nome = 0
        if hn == kh:
            score_nome = 10
        else:
            # 2a) Núcleo do nome (palavras significativas, sem genéricas, sem palavras da cidade)
            #     resolve "PRATA HOTEL" ↔ "HOTEL PRATA"
            nucleo_h = set(p for p in hn.split() if len(p) > 2 and p not in PALAVRAS_GENERICAS and nao_eh_cidade(p))
            nucleo_k = set(p for p in kh.split() if len(p) > 2 and p not in PALAVRAS_GENERICAS and nao_eh_cidade(p))
            if nucleo_h and nucleo_h == nucleo_k:
                score_nome = 10
            elif hn in kh or kh in hn:
                score_nome = 4
            else:
                ph = [p for p in hn.split() if len(p) > 3 and p not in PALAVRAS_GENERICAS and nao_eh_cidade(p)]
                pk = [p for p in kh.split() if len(p) > 3 and p not in PALAVRAS_GENERICAS and nao_eh_cidade(p)]
                if ph and pk:
                    matches  = sum(1 for p in ph if p in kh)
                    matchesk = sum(1 for p in pk if p in hn)
                    score_nome = (matches + matchesk) * 2

        # Sem match de nome → não é o mesmo hotel, mesmo na mesma cidade
        if score_nome == 0:
            continue

        # 3) Bônus de cidade só faz sentido se já houve match de nome
        score = score_nome
        if cidade_bate and cn:
            score += 5

        # Score mínimo dinâmico — desconsiderando palavras da cidade
        palavras_sig = [p for p in hn.split() if len(p) > 3 and p not in PALAVRAS_GENERICAS and nao_eh_cidade(p)]
        score_min = 12 if len(palavras_sig) == 0 else (8 if len(palavras_sig) == 1 else 5)

        if score >= score_min and score > melhor_score:
            melhor_score = score
            melhor = v

    return melhor

def buscar_historico(hotel_name, cidade_str, hotel_hist_sgl, hotel_hist_dbl, hotel_emissores, hotel_pagamento):
    hn = norm(hotel_name)
    cn = norm(cidade_str.split("/")[0].strip())
    best_key, best_score = None, 0
    for k in hotel_emissores:
        parts = k.split("||")
        kh = norm(parts[0])
        kc = norm(parts[1]) if len(parts) > 1 else ""
        score = 0
        if hn in kh or kh in hn:
            score += 2
        if cn in kc or kc in cn:
            score += 1
        if score >= 2 and score > best_score:
            best_score, best_key = score, k
    if not best_key:
        return None
    emissores  = hotel_emissores[best_key]
    t_sgl = hotel_hist_sgl.get(best_key, [])
    t_dbl = hotel_hist_dbl.get(best_key, [])
    t_all = t_sgl + t_dbl

    def stats(lst):
        if not lst: return None, []
        return round(sum(lst)/len(lst)), sorted({round(t) for t in lst})[:5]

    media_sgl, faixa_sgl = stats(t_sgl)
    media_dbl, faixa_dbl = stats(t_dbl)
    media_all, faixa_all = stats(t_all)

    # Forma de pagamento mais frequente
    pgto_counter = hotel_pagamento.get(best_key, Counter())
    forma_pgto_mais_freq = pgto_counter.most_common(1)[0][0] if pgto_counter else ""

    return {
        "dono":        emissores.most_common(1)[0][0],
        "n_emissores": len(emissores),
        "n_reservas":  sum(emissores.values()),
        "tarifa_ref":  media_all,
        "tarifas":     faixa_all,
        "tarifa_sgl":  media_sgl,
        "faixa_sgl":   faixa_sgl,
        "tarifa_dbl":  media_dbl,
        "faixa_dbl":   faixa_dbl,
        "forma_pgto":  forma_pgto_mais_freq,  # ← NOVO
    }

def ler_os_xls(raw_bytes):
    import subprocess, tempfile, csv

    with tempfile.TemporaryDirectory() as tmpdir:
        xls_path = os.path.join(tmpdir, "source.xls")
        with open(xls_path, "wb") as f:
            f.write(raw_bytes)

        result = subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "csv",
             "--outdir", tmpdir, xls_path],
            capture_output=True, text=True, timeout=120
        )
        if result.returncode != 0:
            raise RuntimeError(f"LibreOffice falhou: {result.stderr[:300]}")

        csv_path = xls_path.replace(".xls", ".csv")
        with open(csv_path, encoding="utf-8", errors="replace") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

    print(f"  ✓ OS lidas: {len(rows)} linhas")
    return rows

def limpar_nome_hotel(hotel_raw, cidade_raw):
    """
    Tenta extrair o nome real do hotel quando a consultora preenche o campo
    com cidade-estado-hotel-endereço junto.

    Exemplos:
      "SÃO RAIMUNDO NONATO - PIAUI - MEGA EXPRESS HOTEL II - PRACA CEL MILANEZ, SN CIPO"
        → "MEGA EXPRESS HOTEL II"
      "Hotel Pousada do Sol, Av Brasil 200, Centro, Recife/PE"
        → "Hotel Pousada do Sol"
      "Logic Hoteis Volta Redonda" → "Logic Hoteis Volta Redonda" (sem mudança)

    Estratégia:
      1) Se tem múltiplos " - " (3+), assumir formato CIDADE-ESTADO-HOTEL-ENDEREÇO
         e pegar o segmento que tem palavra-chave de hotel (HOTEL/POUSADA/PALACE/etc.)
      2) Se tem ", " seguido de algo parecido com endereço (RUA, AV, R., etc.),
         cortar antes do endereço.
      3) Caso contrário, retorna como está (com strip).
    """
    if not hotel_raw or not isinstance(hotel_raw, str):
        return hotel_raw
    s = hotel_raw.strip()
    if not s or s.lower() in ("nan", "none"):
        return s

    PALAVRAS_HOTEL = {'HOTEL','POUSADA','PALACE','RESORT','PLAZA','INN','SUITE','SUITES',
                      'FLAT','APART','PARK','LODGE','HOSTEL','MOTEL','HOTEIS','HOTÉIS'}

    def tem_palavra_hotel(seg):
        # Verifica se o segmento contém uma palavra-chave de hotel
        seg_up = seg.upper()
        for p in PALAVRAS_HOTEL:
            if p in seg_up.split() or p in seg_up:
                return True
        return False

    def parece_endereco(seg):
        # Detecta segmento que começa com RUA, AV, R., ROD., TRAVESSA, PRAÇA etc.
        # ou que contém número de endereço (SN, S/N, dígitos+vírgula)
        seg_up = seg.upper().strip()
        prefixos = ('RUA ','R. ','R, ','AV ','AV. ','AVENIDA ','ROD ','ROD. ',
                    'RODOVIA ','TRAVESSA ','TRV ','TRV. ','PRACA ','PRAÇA ','PRAC. ',
                    'AL. ','ALAMEDA ','LARGO ','BR-','BR ','PCA ','PÇA ')
        if any(seg_up.startswith(p) for p in prefixos):
            return True
        # contém "SN", "S/N", número de CEP
        if re.search(r'\b(SN|S/N|S\.N\.?)\b', seg_up):
            return True
        if re.search(r'\d{2}\.?\d{3}-?\d{3}', seg_up):  # CEP
            return True
        return False

    # Normalizar separadores: troca múltiplos espaços, padroniza " - "
    s_norm = re.sub(r'\s+', ' ', s)

    # Estratégia 1: split por " - "
    if s_norm.count(' - ') >= 2:
        partes = [p.strip() for p in s_norm.split(' - ') if p.strip()]
        # Procurar segmento com palavra-chave de hotel, NÃO sendo endereço
        candidatos = [p for p in partes if tem_palavra_hotel(p) and not parece_endereco(p)]
        if candidatos:
            # Pega o primeiro candidato (geralmente o nome do hotel vem antes do endereço)
            return candidatos[0].strip()

    # Estratégia 2: split por vírgula — cortar antes do primeiro segmento que parece endereço
    if ',' in s_norm:
        partes = [p.strip() for p in s_norm.split(',')]
        # Achar índice do primeiro segmento que parece endereço
        idx_endereco = None
        for i, p in enumerate(partes):
            if parece_endereco(p):
                idx_endereco = i
                break
        if idx_endereco is not None and idx_endereco > 0:
            return ', '.join(partes[:idx_endereco]).strip()

    return s_norm


def processar(os_rows, sourcing, hotel_hist_sgl, hotel_hist_dbl, hotel_emissores, hotel_pagamento, os_finalizadas=None):
    """
    Processa OS pendentes do XLS, agrupando por hotel+cidade.
    Se os_finalizadas é fornecido (set de números de OS já emitidas no histórico),
    filtra do output as OS que já estão lá — evita "OS fantasmas" no painel.
    """
    if os_finalizadas is None:
        os_finalizadas = set()

    # Filtrar OS já finalizadas ANTES de agrupar
    os_rows_pendentes = []
    n_filtradas = 0
    os_filtradas_sample = []
    for r in os_rows:
        try:
            n_os = int(float(str(r.get("NÚMERO DA OS", 0)).replace(",", ".")))
        except (ValueError, TypeError):
            n_os = 0
        if n_os and n_os in os_finalizadas:
            n_filtradas += 1
            if len(os_filtradas_sample) < 5:
                os_filtradas_sample.append(n_os)
            continue
        os_rows_pendentes.append(r)

    if n_filtradas > 0:
        sample_str = ", ".join(f"#{n}" for n in os_filtradas_sample)
        if n_filtradas > len(os_filtradas_sample):
            sample_str += f" ...e mais {n_filtradas - len(os_filtradas_sample)}"
        print(f"  OK {n_filtradas} OS j\u00e1 finalizadas no hist\u00f3rico foram removidas da listagem ({sample_str})")

    grupos = defaultdict(list)
    for r in os_rows_pendentes:
        hotel_raw = str(r.get("NOME DO HOTEL", "")).strip()
        cidade = str(r.get("CIDADE", "")).strip()
        if not hotel_raw or hotel_raw.lower() in ("nan", "none", ""):
            hotel = "Sem nome"
        else:
            hotel = limpar_nome_hotel(hotel_raw, cidade)
            if not hotel:
                hotel = "Sem nome"
        if not cidade or cidade.lower() in ("nan", "none", ""):
            cidade = ""
        key = (hotel, cidade)
        grupos[key].append(r)

    grupos_ord = sorted(grupos.items(), key=lambda x: (-len(x[1]), x[0][0]))

    enriched = []
    for (hotel, cidade), rows in grupos_ord:
        os_list = []
        for r in rows:
            try:
                n_os = int(float(str(r.get("NÚMERO DA OS", 0)).replace(",", ".")))
            except Exception:
                n_os = 0
            if not n_os:
                continue

            tarifa_raw = r.get("VALOR DA DIÁRIA", "") or r.get("VALOR DA DIARIA", "")
            try:
                t_val = int(float(str(tarifa_raw).replace(",", ".")))
                if t_val <= 0:
                    t_val = None
            except Exception:
                t_val = None

            status_orig = str(r.get("STATUS DA VIAGEM", ""))
            q = "Cotação" if "Cotação" in status_orig or "Cotacao" in status_orig else "Emissão"

            apto = str(r.get("TIPO DE APARTAMENTO", "Individual")).strip()
            if apto.lower() in ("nan", "none", ""):
                apto = "Individual"

            hosp = str(r.get("NOME DO HÓSPEDE", r.get("NOME DO HOSPEDE", ""))).strip()[:30]
            obs  = str(r.get("OBSERVAÇÃO", r.get("OBSERVACAO", ""))).strip()[:80]
            obs  = obs if obs.lower() not in ("nan", "none") else ""

            os_list.append({
                "n":        n_os,
                "q":        q,
                "cons":     str(r.get("NOME DO CONSULTOR", "")).strip(),
                "t":        t_val,
                "d":        fmt_date(r.get("DATA DE CHECK-IN", "")),
                "hosp":     hosp if hosp.lower() not in ("nan", "none") else "",
                "apto":     apto,
                "obs_orig": obs,
            })

        if not os_list:
            continue

        tarifas_os = [o["t"] for o in os_list if o["t"]]
        src  = buscar_sourcing(hotel, cidade, sourcing)
        hist = buscar_historico(hotel, cidade, hotel_hist_sgl, hotel_hist_dbl,
                                hotel_emissores, hotel_pagamento)

        if src:
            div = False
            for o in os_list:
                if not o["t"]:
                    continue
                ref = src["tarifa_dup"] if o["apto"] == "Duplo" and src.get("tarifa_dup") else src.get("tarifa_ind")
                if ref and abs(o["t"] - ref) > ref * DIV_THRESHOLD:
                    div = True
                    break
            tipo       = "div" if div else "acordo"
            tarifa_ref = src["tarifa_ind"]
        elif hist:
            tipo       = "historico"
            tarifa_ref = hist["tarifa_ref"]
            if len(set(tarifas_os)) > 1 and tarifas_os:
                if max(tarifas_os) - min(tarifas_os) > min(tarifas_os) * 0.10:
                    tipo = "div"
        else:
            tipo       = "sem_ref"
            tarifa_ref = None
            if len(set(tarifas_os)) > 1 and tarifas_os:
                if max(tarifas_os) - min(tarifas_os) > min(tarifas_os) * 0.10:
                    tipo = "div"

        consultores = list({o["cons"] for o in os_list})

        enriched.append({
            "h":    hotel,
            "c":    cidade,
            "tipo": tipo,
            "tr":   tarifa_ref,
            "dono": hist["dono"].title() if hist else "",
            "mc":   len(consultores) > 1,
            "nrh":  hist["n_reservas"] if hist else 0,
            "nem":  hist["n_emissores"] if hist else 0,
            "ht":   hist["tarifas"] if hist else [],
            "ht_sgl": hist.get("faixa_sgl", []) if hist else [],
            "ht_dbl": hist.get("faixa_dbl", []) if hist else [],
            "tr_sgl": hist.get("tarifa_sgl") if hist else None,
            "tr_dbl": hist.get("tarifa_dbl") if hist else None,
            "fp":   hist.get("forma_pgto", "") if hist else "",  # ← NOVO
            "src": {
                "ti":   str(src["tarifa_ind"]) if src and src.get("tarifa_ind") else "",
                "td":   str(src["tarifa_dup"]) if src and src.get("tarifa_dup") else "",
                "tt":   src["tipo_tarifa"][:35] if src and src.get("tipo_tarifa") else "",
                "cond": src["condicao"] if src and src.get("condicao") else "",
                "bk":   src["blackout"] if src and src.get("blackout") else "",
                "em":   src["email"] if src and src.get("email") else "",
                "cafe": src["cafe"] if src and src.get("cafe") else "",
            } if src else None,
            "os": os_list,
        })

    return enriched

def gerar_data_js(enriched, timestamp_str, os_finalizadas=None, planilha_ts=""):
    if os_finalizadas is None:
        os_finalizadas = set()
    # Lista de números de OS já finalizadas (do histórico Vol. Hotelaria 2026).
    # Usado pelo painel para filtrar OS de email que apareçam mesmo já tendo sido emitidas.
    finalizadas_js = "[" + ",".join(str(n) for n in sorted(os_finalizadas)) + "]"
    # PLANILHA_TIMESTAMP = quando o arquivo XLS foi modificado pela última vez (Last-Modified do servidor).
    # DATA_TIMESTAMP = quando o data.js foi gerado pelo gerar_data.py.
    # São diferentes: o XLS pode ter sido atualizado às 14h, mas o data.js só foi regenerado às 16h.
    # Pra Limpar OS órfãs no painel, usar PLANILHA_TIMESTAMP (representa "até quando os dados estão válidos").
    lines = [
        f'// Gerado automaticamente em {timestamp_str}',
        f'const DATA_TIMESTAMP = "{timestamp_str}";',
        f'const PLANILHA_TIMESTAMP = "{planilha_ts}";',
        f'const OS_FINALIZADAS = new Set({finalizadas_js});',
        'const DATA = [',
    ]

    for g in enriched:
        os_items = []
        for o in g["os"]:
            t_val    = str(o["t"]) if o["t"] is not None else "null"
            hosp     = o.get("hosp", "").replace('"', ' ').replace("'", " ")
            obs_orig = o.get("obs_orig", "").replace('"', ' ').replace('\n', ' ')[:70]
            apto     = o.get("apto", "Individual")
            os_items.append(
                f'{{n:{o["n"]},q:"{o["q"]}",cons:"{o["cons"]}",t:{t_val},'
                f'd:"{o["d"]}",hosp:"{hosp}",apto:"{apto}",obs_orig:"{obs_orig}"}}'
            )

        src = g.get("src")
        src_js = "null"
        if src and any(src.values()):
            parts = []
            if src.get("ti"): parts.append(f'ti:{src["ti"]}')
            if src.get("td"): parts.append(f'td:{src["td"]}')
            if src.get("tt"): parts.append(f'tt:"{src["tt"]}"')
            if src.get("cond"): parts.append(f'cond:"{src["cond"]}"')
            if src.get("bk"):
                bk = src["bk"].replace('"', ' ')
                parts.append(f'bk:"{bk}"')
            if src.get("em"): parts.append(f'em:"{src["em"]}"')
            if src.get("cafe"): parts.append(f'cafe:"{src["cafe"]}"')
            if parts:
                src_js = "{" + ",".join(parts) + "}"

        ht      = json.dumps(g.get("ht", []))
        ht_sgl  = json.dumps(g.get("ht_sgl", []))
        ht_dbl  = json.dumps(g.get("ht_dbl", []))
        tr_sgl  = str(g["tr_sgl"]) if g.get("tr_sgl") else "null"
        tr_dbl  = str(g["tr_dbl"]) if g.get("tr_dbl") else "null"
        fp      = g.get("fp", "").replace('"', ' ')
        dono    = g.get("dono", "").replace('"', ' ')
        mc      = "true" if g.get("mc") else "false"
        tr      = str(g["tr"]) if g.get("tr") else "null"
        h       = g["h"].replace('"', ' ').replace("'", ' ')
        c       = g["c"].replace('"', ' ')

        line = (
            f'  {{h:"{h}",c:"{c}",tipo:"{g["tipo"]}",'
            f'tr:{tr},dono:"{dono}",mc:{mc},nrh:{g.get("nrh",0)},nem:{g.get("nem",0)},'
            f'ht:{ht},ht_sgl:{ht_sgl},ht_dbl:{ht_dbl},tr_sgl:{tr_sgl},tr_dbl:{tr_dbl},'
            f'fp:"{fp}",src:{src_js},os:[{",".join(os_items)}]}}'
        )
        lines.append(line + ',')

    lines.append('];')
    return "\n".join(lines)

def main():
    print("=" * 60)
    print("Eurofarma Dashboard — gerador de data.js")
    print("=" * 60)

    if not ONEDRIVE_URL:
        print("ERRO: ONEDRIVE_URL não definida.")
        sys.exit(1)
    if not SOURCING_URL:
        print("ERRO: SOURCING_URL não definida.")
        sys.exit(1)

    print("\n[1/4] Baixando bases de dados...")
    xls_bytes, xls_lm           = download_bytes(ONEDRIVE_URL, "SourceHoteis (OS)")
    sourcing_bytes, sourcing_lm = download_bytes(SOURCING_URL, "Sourcing 2026")

    # Timestamp lógico da planilha = mais ANTIGO dos arquivos baixados.
    # Razão: representa "até que momento os dados estão atualizados".
    # Se um arquivo é de ontem e o outro é de hoje, o conjunto está válido só até ontem.
    planilha_lm = None
    if xls_lm and sourcing_lm:
        planilha_lm = min(xls_lm, sourcing_lm)
    elif xls_lm:
        planilha_lm = xls_lm
    elif sourcing_lm:
        planilha_lm = sourcing_lm
    planilha_ts = planilha_lm.strftime("%d/%m/%Y %H:%M") if planilha_lm else ""

    print("\n[2/4] Processando Sourcing e Histórico...")
    sourcing = carregar_sourcing(sourcing_bytes)
    hotel_hist_sgl, hotel_hist_dbl, hotel_emissores, hotel_pagamento, os_finalizadas = carregar_historico(sourcing_bytes)

    print("\n[3/4] Lendo OS do XLS...")
    os_rows  = ler_os_xls(xls_bytes)
    enriched = processar(os_rows, sourcing, hotel_hist_sgl, hotel_hist_dbl,
                         hotel_emissores, hotel_pagamento, os_finalizadas)

    from collections import Counter as C
    tipos = C(g["tipo"] for g in enriched)
    n_os  = sum(len(g["os"]) for g in enriched)
    print(f"\n  Grupos: {len(enriched)} | OS: {n_os}")
    print(f"  Tipos: {dict(tipos)}")
    print(f"  Multi-consultores: {sum(1 for g in enriched if g['mc'])}")

    print("\n[4/4] Gerando data.js...")
    ts  = datetime.now(TZ_BR).strftime("%d/%m/%Y %H:%M")
    if planilha_ts:
        print(f"  Planilha: {planilha_ts} | Geração: {ts}")
    else:
        print(f"  Geração: {ts} (sem timestamp da planilha)")
    js  = gerar_data_js(enriched, ts, os_finalizadas, planilha_ts)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write(js)
    print(f"  ✓ {OUTPUT_FILE} gravado ({len(js):,} chars)")
    print("\n✅ Concluído com sucesso.")

if __name__ == "__main__":
    main()
